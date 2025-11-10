from ..models.domain import ReportTemplate, Account
from .statement_generation_service import CalculatedData
from typing import List, Dict, Any
from jinja2 import Environment, Template
from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import io
import json

# A basic Jinja2 template as described in the blueprint
PDF_TEMPLATE_STR = """
<html>
<head>
    <style>
        body { font-family: sans-serif; }
        .report-title { font-size: 24px; font-weight: bold; text-align: center; margin-bottom: 20px; }
        .section-title { font-size: 18px; font-weight: bold; margin-top: 15px; border-bottom: 1px solid #333; }
        .row { display: flex; justify-content: space-between; padding: 4px 0; }
        .label { text-indent: 20px; }
        .sub-head { text-indent: 40px; font-style: italic; }
        .value { font-weight: bold; }
        .total { font-weight: bold; border-top: 1px solid #000; padding-top: 5px; }
    </style>
</head>
<body>
    <div class="report-title">{{ report_name }}</div>
    
    {% for item in template_def %}
        {% if item.type == 'section_title' %}
            <div class="section-title">{{ item.label }}</div>
        
        {% elif item.type == 'head' %}
            <div class="row">
                <span class="label">{{ item.label }}</span>
                <span class="value">{{ data.by_head.get(item.account_id, 0) | round(2) }}</span>
            </div>
            
        {% elif item.type == 'sub_head' %}
            <div class="row">
                <span class="sub-head">{{ item.label }}</span>
                <span class="value">{{ data.by_sub_head.get(item.account_id, 0) | round(2) }}</span>
            </div>

        {% elif item.type == 'total' %}
            <div class="row total">
                <span class="label">{{ item.label }}</span>
                <span class="value">{{ data.by_category.get(item.account_id, 0) | round(2) }}</span>
            </div>
        {% endif %}
    {% endfor %}
</body>
</html>
"""

PDF_TEMPLATE = Environment().from_string(PDF_TEMPLATE_STR)

def render_pdf(
    template: ReportTemplate,
    accounts: Dict[int, Account],
    data: CalculatedData
) -> bytes:
    """
    Renders the calculated financial data into a PDF byte stream
    using a report template.
    """
    try:
        template_def = json.loads(template.template_definition)
    except json.JSONDecodeError:
        template_def = [] # Or raise error

    # Construct the context for Jinja2
    context = {
        "report_name": template.name,
        "template_def": template_def,
        "data": data,
        "accounts": accounts
    }
    
    html_string = PDF_TEMPLATE.render(context)
    
    pdf_bytes = HTML(string=html_string).write_pdf()
    return pdf_bytes

def render_excel(
    template: ReportTemplate,
    accounts: Dict[int, Account],
    data: CalculatedData
) -> bytes:
    """
    Renders the calculated financial data into an Excel .xlsx byte stream.
    """
    try:
        template_def = json.loads(template.template_definition)
    except json.JSONDecodeError:
        template_def = []

    wb = Workbook()
    ws = wb.active
    ws.title = template.name[:30]

    # Title
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = template.name
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")

    row_idx = 3
    
    for item in template_def:
        label = item.get("label", "")
        account_id = item.get("account_id")
        
        if item['type'] == 'section_title':
            ws.merge_cells(f"A{row_idx}:D{row_idx}")
            cell = ws[f"A{row_idx}"]
            cell.value = label
            cell.font = Font(bold=True, size=14)
            row_idx += 1
            
        elif item['type'] == 'head':
            ws[f"B{row_idx}"] = label
            cell = ws[f"D{row_idx}"]
            cell.value = data["by_head"].get(account_id, 0)
            cell.font = Font(bold=True)
            cell.number_format = '#,##0.00'
            row_idx += 1
            
        elif item['type'] == 'sub_head':
            ws[f"C{row_idx}"] = label
            cell = ws[f"D{row_idx}"]
            cell.value = data["by_sub_head"].get(account_id, 0)
            cell.number_format = '#,##0.00'
            row_idx += 1

        elif item['type'] == 'total':
            ws[f"B{row_idx}"] = label
            cell = ws[f"D{row_idx}"]
            cell.value = data["by_category"].get(account_id, 0)
            cell.font = Font(bold=True)
            cell.number_format = '#,##0.00'
            cell.border = Border(top=Side(style='thin'))
            row_idx += 1
            
        row_idx += 1 # Add a little space

    # Save to an in-memory stream
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue()